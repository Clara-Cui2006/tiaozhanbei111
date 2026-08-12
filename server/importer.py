from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

REQUIRED_FIELDS = ("案件编号", "案件名称", "业务条线", "案件类别")
SPECIAL_POLITICAL_FIELDS = (
    "序号",
    "案件名称",
    "姓名",
    "性别",
    "特殊身份",
    "涉案地点（非西城区标注具体哪个外区、外省）",
    "是否西城户籍",
    "移送时间",
    "案由",
    "简要案情(一句话）",
)
ALLOWED_STREET_STATUS = {"已确认街道", "待确认地址", "跨街道案件", "与西城无地域关联"}
ALLOWED_TRANSFER_STATUS = {"未形成线索", "系统研判", "人工复核", "研判确认", "内部移送", "办理反馈", "纳入统计"}
ALLOWED_POLITICAL_REVIEW_STATUS = {"不属于政治安全", "待人工复核", "人工研判", "研判确认", "纳入统计"}
ALLOWED_POLITICAL_RISK_LEVEL = {"关注", "低风险", "中风险", "高风险"}

XICHENG_STREETS = (
    "西长安街街道",
    "新街口街道",
    "月坛街道",
    "展览路街道",
    "德胜街道",
    "金融街街道",
    "什刹海街道",
    "大栅栏街道",
    "天桥街道",
    "椿树街道",
    "陶然亭街道",
    "广安门内街道",
    "牛街街道",
    "白纸坊街道",
    "广安门外街道",
)

EMPTY_VALUES = {"", "无", "否", "None", "none", "NULL", "null", "不详"}


@dataclass
class ParsedImport:
    rows: list[dict[str, Any]]
    subjects: list[dict[str, Any]]
    errors: list[dict[str, Any]]
    sha256: str


def _read_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        return [{**dict(row), "__sheet__": "CSV", "__row__": index} for index, row in enumerate(csv.DictReader(io.StringIO(text)), start=2)]
    if filename.lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers = [_normalize_header(value) for value in next(values)]
            for row_number, row in enumerate(values, start=2):
                if not any(_clean(value) for value in row):
                    continue
                rows.append({**dict(zip(headers, row)), "__sheet__": sheet.title, "__row__": row_number})
        return rows
    if filename.lower().endswith(".json"):
        data = json.loads(content.decode("utf-8"))
        if not isinstance(data, list):
            raise ValueError("JSON 顶层必须为数组")
        return [dict(item) for item in data]
    raise ValueError("仅支持 .xlsx、.csv 或 .json 文件")


def _text(row: dict[str, Any], field: str, default: str = "") -> str:
    return _clean(row.get(field) if field in row else default)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\xa0", " ").replace("\u3000", " ").strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[（(](?:必填|选填)[）)]$", "", _clean(value)).strip()


def _normalize_gender(value: Any) -> str:
    gender = _clean(value).lower()
    if gender in {"男", "男性", "男士", "m", "male", "1"}:
        return "男"
    if gender in {"女", "女性", "女士", "f", "female", "2", "0"}:
        return "女"
    return "未知"


def _subject_from_row(row: dict[str, Any], case_number: str) -> dict[str, Any] | None:
    gender_raw = _first_text(row, "性别", "人员性别", "当事人性别", "犯罪嫌疑人性别")
    name = _first_text(row, "姓名（脱敏）", "姓名", "人员姓名", "当事人姓名", "犯罪嫌疑人姓名")
    age_raw = _first_text(row, "年龄", "人员年龄", "当事人年龄", "犯罪嫌疑人年龄")
    if not any((gender_raw, name, age_raw)):
        return None
    try:
        age = int(float(age_raw)) if age_raw else None
    except ValueError:
        age = None
    return {
        "case_number": case_number,
        "name": name or None,
        "age": age,
        "gender": _normalize_gender(gender_raw),
        "occupation": _first_text(row, "职业", "工作单位/所在学校", "工作单位") or None,
        "special_identity": _first_text(row, "特殊身份", "身份", "其他关注身份") or None,
        "is_resident": _normalize_bool(_first_text(row, "是否常住居民", "是否西城户籍")),
        "crime": _first_text(row, "涉及罪名", "罪名", "案由", "移送案由", "涉嫌案由", "审结案由") or None,
        "summary": _first_text(row, "案情摘要", "案件摘要", "简要案情", "简要案情(一句话）", "附注") or None,
    }


def _first_text(row: dict[str, Any], *fields: str, default: str = "") -> str:
    for field in fields:
        value = _text(row, field)
        if value:
            return value
    return default


def _list_text(row: dict[str, Any], field: str) -> list[str]:
    raw = _text(row, field)
    return _split_tags(raw)


def _split_tags(raw: str) -> list[str]:
    if not raw:
        return []
    return [
        item.strip()
        for item in raw.replace("，", ",").replace("、", ",").replace("；", ",").replace(";", ",").replace("\n", ",").split(",")
        if item.strip() and item.strip() not in EMPTY_VALUES
    ]


def _normalize_bool(raw: str) -> bool | None:
    value = _clean(raw).lower()
    if not value or value in EMPTY_VALUES:
        return None
    if value in {"是", "有", "属实", "true", "yes", "y", "1"}:
        return True
    if value in {"否", "无", "false", "no", "n", "0"}:
        return False
    return None


def _is_special_political_row(row: dict[str, Any]) -> bool:
    headers = {key for key in row.keys() if not key.startswith("__")}
    matched = sum(1 for field in SPECIAL_POLITICAL_FIELDS if field in headers)
    return matched >= 7 and bool(_first_text(row, "案件名称", "案由"))


def _generated_special_case_number(row: dict[str, Any], sheet: str, number: int) -> str:
    source = "|".join([
        sheet,
        _first_text(row, "序号"),
        _first_text(row, "案件名称"),
        _first_text(row, "姓名"),
        _first_text(row, "移送时间"),
        _first_text(row, "案由"),
        str(number),
    ])
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:12].upper()
    return f"PS-{digest}"


def _has_value(row: dict[str, Any], field: str) -> bool:
    return _text(row, field) not in EMPTY_VALUES


def _is_yes(row: dict[str, Any], field: str) -> bool:
    return _text(row, field) == "是"


def _append_unique(items: list[str], *values: str) -> list[str]:
    for value in values:
        for item in _split_tags(value):
            if item not in items:
                items.append(item)
    return items


def _extract_street_names(address: str) -> list[str]:
    normalized = address.replace(" ", "")
    names: list[str] = []
    for street in XICHENG_STREETS:
        short = street.removesuffix("街道")
        if street in normalized or f"{short}街道" in normalized:
            names.append(street)
    return names


def _street_fields(row: dict[str, Any], address: str) -> tuple[str, str | None]:
    street_status = _text(row, "街道归属状态")
    street_name = _text(row, "所属街道")
    if street_status:
        return street_status, street_name or None
    streets = _extract_street_names(address)
    if len(streets) == 1:
        return "已确认街道", streets[0]
    if len(streets) > 1:
        return "跨街道案件", "、".join(streets)
    if "西城区" in address or "西城" in address:
        return "待确认地址", None
    return "与西城无地域关联", None


def _governance_themes(row: dict[str, Any]) -> list[str]:
    themes = _list_text(row, "治理主题标签")
    _append_unique(themes, _text(row, "专项活动名称"), _text(row, "专项活动重点工作"), _text(row, "案件性质"), _text(row, "专业化类型"))
    if _is_yes(row, "涉政治安全类案件") or _is_special_political_row(row):
        _append_unique(themes, "政治安全线索")
    _append_unique(themes, _text(row, "特殊身份"))
    if _is_yes(row, "是否涉外案件"):
        _append_unique(themes, "涉外风险")
    if _is_yes(row, "是否涉及单位犯罪"):
        _append_unique(themes, "单位犯罪")
    if _is_yes(row, "是否利用电信或网络实施犯罪"):
        _append_unique(themes, "网络犯罪")
    if _is_yes(row, "犯罪案件是否涉及养老诈骗"):
        _append_unique(themes, "养老诈骗")
    if _is_yes(row, "是否涉及利用未成年人犯罪") or _is_yes(row, "是否涉嫌性侵未成年人犯罪案件") or "未检" in _text(row, "案件类别"):
        _append_unique(themes, "未成年人保护")
    if _is_yes(row, "案件是否涉及三大攻坚战"):
        _append_unique(themes, "三大攻坚")
    if _is_yes(row, "犯罪案件是否涉及黑社会性质组织") or _is_yes(row, "犯罪案件是否涉及恶势力"):
        _append_unique(themes, "黑恶风险")
    if _has_value(row, "防范化解金融风险关注情形"):
        _append_unique(themes, "金融风险")
    if _has_value(row, "污染防治关注情形"):
        _append_unique(themes, "污染防治")
    return themes


def _key_groups(row: dict[str, Any]) -> list[str]:
    groups: list[str] = []
    for field in ("重点人群标签", "特殊群体", "身份", "其他关注身份", "被害人情况", "当事人类型"):
        _append_unique(groups, _text(row, field))
    if _is_yes(row, "是否党委政府领导"):
        _append_unique(groups, "党委政府领导")
    if _is_yes(row, "是否乡镇基层组织人员") or _is_yes(row, "是否村级组织人员"):
        _append_unique(groups, "基层组织人员")
    if _is_yes(row, "是否残疾人"):
        _append_unique(groups, "残疾人")
    if _is_yes(row, "是否为单位"):
        _append_unique(groups, "单位主体")
    if _normalize_bool(_first_text(row, "是否西城户籍")) is True:
        _append_unique(groups, "西城户籍")
    return groups


def _key_industries(row: dict[str, Any]) -> list[str]:
    industries = _list_text(row, "重点行业标签")
    for field in ("受侵犯企业类型", "移送案件的行政执法机关类别", "案件性质", "工作单位/所在学校"):
        _append_unique(industries, _text(row, field))
    return industries


def _political_fields(row: dict[str, Any], legal_cause: str, street_name: str | None, address: str) -> tuple[str, str, str, str, str, str, str]:
    is_political = _is_yes(row, "涉政治安全类案件") or _is_special_political_row(row)
    is_foreign = _is_yes(row, "是否涉外案件")
    review_status = _text(row, "政治安全研判状态") or ("待人工复核" if is_political else "不属于政治安全")
    risk_level = _text(row, "政治安全风险等级") or ("关注" if is_political else "")
    topic = _text(row, "政治安全专题") or _text(row, "特殊身份")
    if not topic:
        if is_political:
            topic = "特殊案件线索" if _is_special_political_row(row) else "政治安全线索"
        elif is_foreign:
            topic = "涉外风险"
    behavior = _text(row, "行为内容") or _text(row, "案由") or (legal_cause if is_political else "")
    subject = _text(row, "涉及主体") or _text(row, "特殊身份")
    if not subject and is_political:
        if is_foreign:
            subject = "涉外关联人员"
        elif _is_yes(row, "是否为单位"):
            subject = "单位主体"
        elif _is_yes(row, "是否党委政府领导"):
            subject = "重点岗位人员"
        else:
            subject = "重点关注人员"
    location = _text(row, "地点因素") or street_name or address
    spread = _text(row, "传播影响")
    return topic, location, behavior, subject, spread, review_status, risk_level


def parse_import(filename: str, content: bytes) -> ParsedImport:
    rows = _read_rows(filename, content)
    normalized: list[dict[str, Any]] = []
    subjects: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows, start=2):
        number = int(row.get("__row__") or index)
        sheet = _text(row, "__sheet__")
        is_special_political = _is_special_political_row(row)
        case_number = _first_text(row, "案件编号", "部门受案号", "统一受案号")
        related_case_number = _first_text(row, "关联案号", "关联案件编号", "案号")
        # “涉案人员信息”等独立工作表通过关联案号连接案件，不再被误判为案件行。
        if related_case_number and not case_number and not _first_text(row, "案件名称"):
            subject = _subject_from_row(row, related_case_number)
            if subject:
                subjects.append(subject)
            continue
        case_number = case_number or related_case_number
        if not case_number and is_special_political:
            case_number = _generated_special_case_number(row, sheet, number)
        case_name = _first_text(row, "案件名称")
        department = _first_text(row, "业务条线", "承办部门") or ("政治安全专项" if is_special_political else "")
        category = _first_text(row, "案件类别") or ("政治安全特殊案件" if is_special_political else "")
        missing = []
        if not case_number:
            missing.append("部门受案号/案件编号")
        if not case_name:
            missing.append("案件名称")
        if not department:
            missing.append("承办部门/业务条线")
        if not category:
            missing.append("案件类别")
        if missing:
            errors.append({"row": number, "field": ",".join(missing), "message": f"{sheet} 必填字段缺失"})
            continue
        if case_number in seen:
            errors.append({"row": number, "field": "部门受案号/案件编号", "message": f"{sheet} 文件内案件编号重复"})
            continue
        seen.add(case_number)
        address = _first_text(row, "地址", "涉案地点（非西城区标注具体哪个外区、外省）", "主要作案地", "住所地详细地址", "工作单位及地址", "联系地址")
        street_status, street_name = _street_fields(row, address)
        if street_status not in ALLOWED_STREET_STATUS:
            errors.append({"row": number, "field": "街道归属状态", "message": "街道状态不符合规定口径"})
            continue
        if street_status == "已确认街道" and not street_name:
            errors.append({"row": number, "field": "所属街道", "message": "已确认街道的数据必须填写唯一所属街道"})
            continue
        transfer_status = _text(row, "内部线索状态", "未形成线索")
        if not _text(row, "内部线索状态") and "线索" in _text(row, "附注"):
            transfer_status = "内部移送"
        if transfer_status not in ALLOWED_TRANSFER_STATUS:
            errors.append({"row": number, "field": "内部线索状态", "message": "内部线索状态不符合闭环口径"})
            continue
        crime = _first_text(row, "罪名", "案由", "移送案由", "涉嫌案由", "审结案由")
        legal_cause = _first_text(row, "法定罪名/案由", "案由", "移送案由（细分）", "涉嫌案由（细分）", "审结案由（细分）", "移送案由", "涉嫌案由", "审结案由", "案件性质", default=crime or category)
        political_topic, political_location_factor, political_behavior_content, political_subject, political_spread_impact, political_review_status, political_risk_level = _political_fields(row, legal_cause, street_name, address)
        if political_review_status not in ALLOWED_POLITICAL_REVIEW_STATUS:
            errors.append({"row": number, "field": "政治安全研判状态", "message": "政治安全研判状态不符合人工复核口径"})
            continue
        if political_risk_level and political_risk_level not in ALLOWED_POLITICAL_RISK_LEVEL:
            errors.append({"row": number, "field": "政治安全风险等级", "message": "政治安全风险等级不符合规定口径"})
            continue
        normalized.append({
            "case_number": case_number,
            "case_name": case_name,
            "department": department,
            "category": category,
            "crime": crime,
            "legal_cause": legal_cause,
            "governance_themes": _governance_themes(row),
            "accepted_date": _first_text(row, "受理日期", "移送时间", "检察机关受理日期", "民事部门收案日期"),
            "closed_date": _first_text(row, "办结日期", "审结日期", "结案日期", "全案_审结日期", "诉前终结案件日期", "审查起诉终结案件日期"),
            "status": _first_text(row, "审结处理结果", "结案情况", "全案_审结处理结果", "审查起诉结果", "案件状态"),
            "street_status": street_status,
            "street_name": street_name,
            "address": address,
            "keywords": "、".join(_governance_themes(row)[:8]),
            "summary": _first_text(row, "案件摘要", "简要案情", "简要案情(一句话）", "附注"),
            "key_groups": _key_groups(row),
            "key_industries": _key_industries(row),
            "internal_transfer_status": transfer_status,
            "prosecutorial_track": _first_text(row, "承办检察条线", "承办检察官", "办案团队"),
            "political_topic": political_topic,
            "political_location_factor": political_location_factor,
            "political_behavior_content": political_behavior_content,
            "political_subject": political_subject,
            "political_spread_impact": political_spread_impact,
            "political_review_status": political_review_status,
            "political_risk_level": political_risk_level,
        })
        subject = _subject_from_row(row, case_number)
        if subject:
            subjects.append(subject)

    imported_case_numbers = {row["case_number"] for row in normalized}
    for subject in subjects:
        if subject["case_number"] not in imported_case_numbers:
            errors.append({"row": 0, "field": "关联案号", "message": f"涉案人员未找到关联案件：{subject['case_number']}"})
    subjects = [subject for subject in subjects if subject["case_number"] in imported_case_numbers]
    return ParsedImport(normalized, subjects, errors, hashlib.sha256(content).hexdigest())
